import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pyfiglet
import os
import argparse
import sys

def normalizza_orario(orario):
    return orario.replace('.', ':')

def parse_data(data_str):
    try:
        return datetime.datetime.strptime(data_str, "%d/%m/%Y")
    except ValueError:
        print(f"Formato data non valido: {data_str}. Usa il formato dd/mm/yyyy.")
        sys.exit(1)

def inserisci_settimana(args):
    if args.data_inizio and args.data_fine:
        data_inizio = parse_data(args.data_inizio)
        data_fine = parse_data(args.data_fine)
    else:
        while True:
            data_inizio = parse_data(input("Inserisci la data di inizio della settimana (dd/mm/yyyy): "))
            data_fine = parse_data(input("Inserisci la data di fine della settimana (dd/mm/yyyy): "))
            if data_inizio > data_fine:
                print("La data di inizio deve essere precedente alla data di fine. Riprova.")
            else:
                break
    return data_inizio, data_fine

def inserisci_orario_lavorativo(giorno_settimana):
    while True:
        orario_inizio = input(f"Inserisci Orario di Inizio Lavoro per {giorno_settimana} (HH:MM o HH.MM) o '/' se non lavori: ")
        if orario_inizio.strip() == '/':
            return None, None
        try:
            orario_fine = input(f"Inserisci Orario di Fine Lavoro per {giorno_settimana} (HH:MM o HH.MM): ")
            
            orario_inizio = datetime.datetime.strptime(normalizza_orario(orario_inizio), "%H:%M")
            orario_fine = datetime.datetime.strptime(normalizza_orario(orario_fine), "%H:%M")
            
            if orario_fine < orario_inizio:
                print("L'orario di fine deve essere successivo all'orario di inizio. Riprova.")
            else:
                return orario_inizio, orario_fine
        except ValueError:
            print("Formato orario non valido. Riprova.")

def calcola_straordinari(ore_lavorate, ore_standard=8):
    return max(0, ore_lavorate - ore_standard)

def applica_stile_excel(writer, df):
    workbook = writer.book
    worksheet = writer.sheets['Ore Lavorative']
    
    # Stile per il titolo
    title_cell = worksheet['A1']
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Stile per l'intestazione
    header_cells = worksheet['A2:E2']
    for cell in header_cells[0]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    # Autofit delle colonne
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    parser = argparse.ArgumentParser(description="HourWorker: Calcola le ore lavorate settimanalmente")
    parser.add_argument("--data_inizio", help="Data di inizio settimana (dd/mm/yyyy)")
    parser.add_argument("--data_fine", help="Data di fine settimana (dd/mm/yyyy)")
    parser.add_argument("--output", help="Percorso del file di output Excel")
    args = parser.parse_args()

    # Titolo
    titolo = pyfiglet.figlet_format("HourWorker", font="slant")
    print(titolo)
    print("_" * 60)

    # Cartella di destinazione per il file Excel
    cartella_destinazione = args.output if args.output else r"C:\Users\batte\Desktop\HourWorker"
    if not os.path.exists(cartella_destinazione):
        os.makedirs(cartella_destinazione)

    data_inizio, data_fine = inserisci_settimana(args)
    giorni_settimana = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
    dati_settimanali = []

    for giorno_settimana in giorni_settimana:
        orario_inizio, orario_fine = inserisci_orario_lavorativo(giorno_settimana)
        if orario_inizio is None:
            print(f"Non lavori il {giorno_settimana}.")
            dati_settimanali.append((giorno_settimana, data_inizio.strftime("%d/%m/%Y"), 0, 0, 0))
        else:
            ore_lavorate = (orario_fine - orario_inizio).seconds / 3600
            ore_straordinario = calcola_straordinari(ore_lavorate)
            dati_settimanali.append((giorno_settimana, data_inizio.strftime("%d/%m/%Y"), ore_lavorate, ore_straordinario, ore_lavorate - ore_straordinario))
            print(f"Hai lavorato {ore_lavorate:.2f} ore il {giorno_settimana} {data_inizio.strftime('%d/%m/%Y')} (di cui {ore_straordinario:.2f} ore di straordinario)")
        data_inizio += datetime.timedelta(days=1)

    df = pd.DataFrame(dati_settimanali, columns=['Giorno', 'Data', 'Ore Totali', 'Ore Straordinario', 'Ore Ordinarie'])
    somma_ore = df['Ore Totali'].sum()
    somma_straordinari = df['Ore Straordinario'].sum()
    somma_ordinarie = df['Ore Ordinarie'].sum()
    print(f"Totale ore lavorate nella settimana: {somma_ore:.2f}")
    print(f"Di cui ore straordinarie: {somma_straordinari:.2f}")
    print(f"E ore ordinarie: {somma_ordinarie:.2f}")

    df.loc[len(df.index)] = ['Totale', '', somma_ore, somma_straordinari, somma_ordinarie]  # Aggiunge la riga del totale

    titolo_settimana = f"Settimana dal {data_inizio.strftime('%d/%m/%Y')} al {data_fine.strftime('%d/%m/%Y')}"

    # Crea il file Excel nella cartella specificata
    file_excel_path = os.path.join(cartella_destinazione, 'ore_lavorate_settimana.xlsx')
    writer = pd.ExcelWriter(file_excel_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Ore Lavorative', startrow=1)  # Inizia dalla seconda riga per lasciare spazio al titolo

    # Aggiunge il titolo della settimana
    workbook = writer.book
    worksheet = writer.sheets['Ore Lavorative']
    worksheet.merge_cells('A1:E1')
    worksheet['A1'] = titolo_settimana

    applica_stile_excel(writer, df)

    writer.save()
    print(f"File Excel creato: {file_excel_path}")

if __name__ == "__main__":
    main()
